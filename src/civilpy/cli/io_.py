#  CivilPy
#  Copyright (C) 2019-2026 Dane Parks
#
#  SPDX-License-Identifier: MIT
#  See the LICENSE file in the project root for full license text.

"""Result model and file writers for the CivilPy CLI.

Every command returns a :class:`CommandResult`: one or more
:class:`ResultTable` (columns carry their unit; rows hold raw Python
values so files get numbers, not strings) plus provenance — the one-shot
command line, the inputs, and a hash of each input file — so a workbook
can always be traced back to what produced it.

Writers dispatch on the output suffix: ``.xlsx`` is one workbook, a sheet
per table plus a ``provenance`` sheet; ``.csv`` is one file per table
(suffixed with the table name when there is more than one).
"""

from __future__ import annotations

import csv
import datetime as _dt
import hashlib
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, List, Optional, Tuple


@dataclass(frozen=True)
class Column:
    """One output column: ``unit`` renders dimmed in headers and as a
    ``(unit)`` suffix in files; ``fmt`` is a terminal-only format spec
    (files keep raw values)."""

    name: str
    unit: Optional[str] = None
    fmt: Optional[str] = None

    @property
    def header(self) -> str:
        return f"{self.name} ({self.unit})" if self.unit else self.name


@dataclass
class ResultTable:
    title: str
    columns: List[Column]
    rows: List[Tuple[Any, ...]]
    notes: List[str] = field(default_factory=list)


@dataclass
class CommandResult:
    """Tables plus the provenance that lets an output file be reproduced.

    ``exit_code`` lets a command that *completed* still fail the process —
    a validation run reports its errors as tables and exits nonzero so CI
    can gate on it.
    """

    tables: List[ResultTable]
    command: str = ""
    inputs: dict = field(default_factory=dict)
    input_files: List[str] = field(default_factory=list)
    exit_code: int = 0

    def provenance_rows(self) -> List[Tuple[str, str]]:
        rows = [
            ("generated", _dt.datetime.now().isoformat(timespec="seconds")),
            ("civilpy", _version()),
            ("command", self.command),
        ]
        rows += [(k, repr(v)) for k, v in sorted(self.inputs.items())]
        for f in self.input_files:
            rows.append((f"input file", f"{f} (sha256 {_sha256(f)})"))
        return rows


def _version() -> str:
    try:
        from importlib.metadata import version

        return version("civilpy")
    except Exception:  # pragma: no cover - metadata missing in odd installs
        return "unknown"


def _sha256(path: str, chars: int = 12) -> str:
    try:
        digest = hashlib.sha256(Path(path).read_bytes()).hexdigest()
        return digest[:chars]
    except OSError:  # pragma: no cover - file vanished after the run
        return "unreadable"


def _safe_name(title: str, maxlen: int = 31) -> str:
    """Excel sheet titles: max 31 chars, no ``[]:*?/\\``."""
    name = re.sub(r"[\[\]:*?/\\]", "-", title).strip()
    return name[:maxlen] or "Sheet"


def write_result(result: CommandResult, out_path: Path) -> List[Path]:
    """Write ``result`` to ``out_path`` (suffix picks the writer) and
    return the paths written."""
    out_path = Path(out_path)
    suffix = out_path.suffix.lower()
    if suffix == ".xlsx":
        return _write_xlsx(result, out_path)
    if suffix == ".csv":
        return _write_csv(result, out_path)
    raise ValueError(
        f"unsupported output type {suffix!r}: use a .csv or .xlsx path"
    )


def _write_csv(result: CommandResult, out_path: Path) -> List[Path]:
    written = []
    multi = len(result.tables) > 1
    for table in result.tables:
        path = out_path
        if multi:
            slug = re.sub(r"\W+", "_", table.title.lower()).strip("_")
            path = out_path.with_name(f"{out_path.stem}_{slug}.csv")
        with open(path, "w", newline="", encoding="utf-8") as fh:
            writer = csv.writer(fh)
            writer.writerow([c.header for c in table.columns])
            writer.writerows([_file_cell(v) for v in row] for row in table.rows)
        written.append(path)
    return written


def _file_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "yes" if value else "no"
    return value


def _write_xlsx(result: CommandResult, out_path: Path) -> List[Path]:
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True)

    for table in result.tables:
        ws = wb.create_sheet(_safe_name(table.title))
        ws.append([c.header for c in table.columns])
        for cell in ws[1]:
            cell.font = header_font
        for row in table.rows:
            ws.append([_file_cell(v) for v in row])
        for note in table.notes:
            ws.append([])
            ws.append([f"Note: {note}"])
        ws.freeze_panes = "A2"
        for i, col in enumerate(table.columns, start=1):
            width = max(
                len(col.header),
                *(len(str(_file_cell(r[i - 1]))) for r in table.rows or [("",)]),
            )
            ws.column_dimensions[get_column_letter(i)].width = min(width + 2, 40)

    ws = wb.create_sheet("provenance")
    ws.append(["field", "value"])
    for cell in ws[1]:
        cell.font = header_font
    for key, value in result.provenance_rows():
        ws.append([key, value])
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 80

    wb.save(out_path)
    return [out_path]
