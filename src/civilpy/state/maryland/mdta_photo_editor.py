"""
CivilPy
Copyright (C) 2019 - Dane Parks

This program is free software: you can redistribute it and/or modify
it under the terms of the GNU Affero General Public License as published by
the Free Software Foundation, either version 3 of the License, or
(at your option) any later version.

This program is distributed in the hope that it will be useful,
but WITHOUT ANY WARRANTY; without even the implied warranty of
MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
GNU Affero General Public License for more details.

You should have received a copy of the GNU Affero General Public License
along with this program.  If not, see <http://www.gnu.org/licenses/>.
"""

import FreeSimpleGUI as sg
from pathlib import Path
import os
import re
from io import BytesIO
from PIL import Image
import pandas as pd
from openpyxl import Workbook

from photos import (convert_filenames_from_excel,
                                    resize_image,
                                    add_timestamp,
                                    get_photos_from_file_list,
                                    get_photo_creation_date)

test_init = True
button_color = ("white", "blue")

# def resize_image(folder):
#     convert_filenames_from_excel(folder)
#     print("converted Images")

def generate_excel(folder, extension):
    path = Path(folder)
    wb = Workbook()
    ws = wb.active

    if folder != "" and path.is_dir():
        files = list(path.glob("*.*"))
        lst = [
            file.name
            for file in files
            if file.suffix.lower() in extension
               and file.is_file()
        ]

    max_length = 0

    for i, file in enumerate(lst):
        if len(str(file)) > max_length:
            max_length = len(str(file))
            ws[f'A{i + 1}'] = str(file)
        else:
            ws[f'A{i + 1}'] = str(file)

    ws.column_dimensions["A"].width = max_length + 2
    ws.column_dimensions["B"].width = max_length + 2

    output_file = path / 'photo_names.xlsx'
    wb.save(output_file)
    os.startfile(output_file)




def update_listbox(listbox_element, folder, extension, substring):
    path = Path(folder)
    filter_ = substring.lower()
    lst = []
    if folder != "" and path.is_dir():
        files = list(path.glob("*.*"))
        lst = [
            file.name
            for file in files
            if file.suffix.lower() in extension
            and filter_ in str(file).lower()
            and file.is_file()
        ]
    listbox_element.update(lst)


def update_image(image_element, filename):
    im = Image.open(filename)
    w, h = size_of_image
    scale = max(im.width / w, im.height / h)
    if scale <= 1:
        image_element.update(filename=filename)
    else:
        im = im.resize(
            (int(im.width / scale), int(im.height / scale)), resample=Image.BICUBIC
        )
        with BytesIO() as output:
            im.save(output, format="PNG")
            data = output.getvalue()
        image_element.update(data=data)


sg.theme("Dark")
sg.set_options(font=("Roboto Mono", 11))

w, h = size_of_image = (700, 600)

layout_top = [
    [
        sg.Text("Photo Folder:"),
        sg.InputText(enable_events=True, key="-FOLDER-"),
        sg.FolderBrowse("Browse", button_color=button_color, size=(7, 1), enable_events=True),
        sg.Button("Generate Excel", key="GENERATE_EXCEL", button_color=button_color, size=(12, 1))
    ],
    [
        sg.Text("Filter:"),
        sg.InputText(enable_events=True, key="-FILTER-"),
        sg.Button("Search", button_color=button_color, size=(7, 1)),
        sg.Button("Resize Images", key="RESIZE_IMAGE", button_color=button_color, size=(12, 1)),
        sg.Checkbox("Generate Timestamps", key="-TimeStamp-")
    ]
]

layout_bottom = [
    [
        sg.Listbox(
            [],
            size=(52, 30),
            enable_events=True,
            select_mode=sg.LISTBOX_SELECT_MODE_SINGLE,
            sbar_background_color="#707070",
            key="-LISTBOX-",
            expand_x=True,
            expand_y=True,
            horizontal_scroll=True,
        )
    ],
]

layout_left = [
    [sg.Column(layout_top, pad=(0, 0), expand_x=True, expand_y=True)],
    [sg.Column(layout_bottom, pad=(0, 0), expand_x=True, expand_y=True)],
]

layout_right = [[sg.Image(background_color="gray", key="-IMAGE-")]]

layout = [
    [
        sg.Column(
            layout_left,
            expand_x=True,
            expand_y=True
        ),
        sg.Column(
            layout_right,
            pad=(0, 0),
            size=(w + 15, h + 15),
            background_color="blue",
            key="-COLUMN-",
        ),
    ],
]

window = sg.Window("MDTA Photo Editor", layout, resizable=True, finalize=True)
window["-IMAGE-"].Widget.pack(fill="both", expand=True)
window["-IMAGE-"].Widget.master.pack(fill="both", expand=True)
window["-IMAGE-"].Widget.master.master.pack(fill="both", expand=True)
window["-COLUMN-"].Widget.pack_propagate(False)

while True:
    event, values = window.read()
    if event == sg.WINDOW_CLOSED:
        break
    # print(event, values)
    if event in ("-FOLDER-", "-FILTER-", "Search"):
        update_listbox(
            window["-LISTBOX-"],
            values["-FOLDER-"],
            (".png", ".gif", ".jpg", ".tif"),
            values["-FILTER-"],
        )
    elif event == "GENERATE_EXCEL":
        generate_excel(
            values["-FOLDER-"],
            (".png", ".gif", ".jpg", ".tif"),
        )
    elif event == "RESIZE_IMAGE":
        # Check if the folder already exists and provide the user a warning
        if os.path.isdir(Path(values["-FOLDER-"]) / 'Renamed_Photos'):
            sg.popup(
                "Warning: The resized folder already exists, delete it before rerunning the command",
                title="File Warning",
                button_color=("white", "red")
            )

        # If renamed photo folder doesn't exist, create it and attempt the conversion
        else:
            os.mkdir(Path(values["-FOLDER-"]) / 'Renamed_Photos')

            file_list = os.listdir(values["-FOLDER-"])
            images = get_photos_from_file_list(file_list)
            excel_file = [
                file
                for file in file_list
                if re.search("(xlsx|xls|xlsm)$", str(file), re.IGNORECASE)
            ][0]
            photo_list_excel = pd.read_excel(Path(values["-FOLDER-"]) / excel_file, 0, header=None)
            file_list = photo_list_excel[0].tolist()
            try:
                new_names = photo_list_excel[1].tolist()
            except KeyError as e:
                new_names = photo_list_excel[0].tolist()

            for file in images:
                # Set the new name of the image based on the excel file
                new_name = new_names[file_list.index(file)]
                image = Image.open(Path(values["-FOLDER-"]) / file)

                if values['-TimeStamp-']:
                    creation_date = get_photo_creation_date(image)
                    image = resize_image(image)
                    image = add_timestamp(image, creation_date)
                else:
                    image = resize_image(image)


                # Check if the value in excel is empty. comes in as a float but checking for NaN crashes program
                if type(new_name)==float:
                    image.save(Path(values["-FOLDER-"]) / 'Renamed_Photos' / Path(file).name)
                else:
                    image.save(Path(values["-FOLDER-"]) / 'Renamed_Photos' / f"{new_name}{Path(file).suffix}")

    elif event == "-LISTBOX-":
        lst = values["-LISTBOX-"]

        if lst:
            update_image(window["-IMAGE-"], Path(values["-FOLDER-"], values["-LISTBOX-"][0]))

window.close()
